import sys
import json
import queue
import threading
import os
import zipfile
import requests
from pathlib import Path
from PyQt5.QtWidgets import (
    QApplication, QWidget, QPushButton, QTextEdit, QGridLayout,
    QHBoxLayout, QVBoxLayout, QDesktopWidget, QLabel, QMenuBar, QScrollArea,
    QAction, QFileDialog, QMessageBox, QTableWidget, QTableWidgetItem,
    QHeaderView, QActionGroup, QDialog, QProgressDialog, QProgressBar
)
from PyQt5.QtCore import Qt, QTimer, QThread, pyqtSignal
from PyQt5.QtGui import QKeyEvent, QFont
import openpyxl
from openpyxl.styles import Font as ExcelFont, Alignment
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
import numpy as np
try:
    import pyaudio
    import vosk
    vosk.SetLogLevel(-1)
    VOSK_AVAILABLE = True
except ImportError:
    VOSK_AVAILABLE = False


class ModelDownloader(QThread):
    """Thread pour télécharger et extraire les modèles Vosk."""
    progress = pyqtSignal(int)
    status = pyqtSignal(str)
    finished = pyqtSignal(bool, str)

    def __init__(self, model_name, model_url):
        super().__init__()
        self.model_name = model_name
        self.model_url = model_url
        self.cancelled = False

    def run(self):
        try:
            appdata_path = Path(os.getenv('APPDATA')) / 'Calculator' / 'vosk-models'
            appdata_path.mkdir(parents=True, exist_ok=True)
            zip_path = appdata_path / f"{self.model_name}.zip"
            extract_path = appdata_path / self.model_name
            self.status.emit(f"Téléchargement de {self.model_name}...")
            response = requests.get(self.model_url, stream=True)
            total_size = int(response.headers.get('content-length', 0))
            downloaded = 0
            with open(zip_path, 'wb') as f:
                for chunk in response.iter_content(chunk_size=8192):
                    if self.cancelled:
                        self.finished.emit(False, "Téléchargement annulé")
                        return
                    if chunk:
                        f.write(chunk)
                        downloaded += len(chunk)
                        if total_size > 0:
                            progress = int((downloaded / total_size) * 100)
                            self.progress.emit(progress)
            self.status.emit(f"Extraction de {self.model_name}...")
            self.progress.emit(0)
            with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                members = zip_ref.namelist()
                total_files = len(members)
                for i, member in enumerate(members):
                    if self.cancelled:
                        self.finished.emit(False, "Extraction annulée")
                        return
                    zip_ref.extract(member, appdata_path)
                    progress = int((i + 1) / total_files * 100)
                    self.progress.emit(progress)
            zip_path.unlink()
            self.finished.emit(True, str(extract_path))
        except Exception as e:
            self.finished.emit(False, f"Erreur: {str(e)}")

    def cancel(self):
        self.cancelled = True


class Calculatrice(QWidget):
    def __init__(self):
        super().__init__()
        self.total = 0
        self.current_number = ""
        self.historique = []
        self.history_of_totals = []
        self.current_language = "fr"

        # ── Undo / Redo stacks ──────────────────────────────────────────────
        # Each state snapshot: {'historique': [...], 'total': float,
        #                        'history_of_totals': [...], 'current_number': str}
        self._undo_stack = []
        self._redo_stack = []

        # Voice recognition variables
        self.is_listening = False
        self.audio_queue = queue.Queue()
        self.vosk_model = None
        self.audio_stream = None
        self.audio_thread = None
        self.current_model_path = None
        self.download_thread = None
        self.pending_language_change = None

        self.language_models = {
            "fr": "vosk-model-small-fr-0.22",
            "ar": "vosk-model-small-ar-tn-0.1-linto",
            "en": "vosk-model-small-en-us-0.15"
        }
        self.model_urls = {
            "vosk-model-small-ar-tn-0.1-linto": "https://alphacephei.com/vosk/models/vosk-model-small-ar-tn-0.1-linto.zip",
            "vosk-model-ar-mgb2-0.4": "https://alphacephei.com/vosk/models/vosk-model-ar-mgb2-0.4.zip",
            "vosk-model-small-fr-0.22": "https://alphacephei.com/vosk/models/vosk-model-small-fr-0.22.zip",
            "vosk-model-small-en-us-0.15": "https://alphacephei.com/vosk/models/vosk-model-small-en-us-0.15.zip"
        }

        self.translations = {
            "fr": {
                "title": "Score Calculator",
                "file_menu": "Fichier",
                "language_menu": "Langue",
                "help_menu": "Aide",
                "shortcuts_action": "Raccourcis & Manipulations",
                "export": "Exporter vers Excel",
                "about": "À propos",
                "quit": "Quitter",
                "french": "Français",
                "arabic": "العربية",
                "english": "English",
                "clear_numbers": "Effacer Nombres",
                "clear_totals": "Effacer Totaux",
                "current_input": "Saisie en cours:",
                "counter_header": "#",
                "value_header": "Valeur",
                "total_header": "Total",
                "tooltip_history": "Historique des nombres (double-clic pour modifier)",
                "tooltip_totals": "Historique des totaux enregistrés",
                "tooltip_clear_numbers": "Efface la liste des nombres ajoutés (Échap)",
                "tooltip_clear_totals": "Efface l'historique des totaux (Suppr)",
                "tooltip_total": "Total courant (Insert ou Ctrl+0 pour enregistrer)",
                "tooltip_enter": "Valider le nombre (Entrée)",
                "tooltip_backspace": "Effacer (Retour arrière)",
                "export_success": "Export réussi",
                "export_success_msg": "Les données ont été exportées vers:\n{0}",
                "export_error": "Erreur",
                "export_error_msg": "Erreur lors de l'export:\n{0}",
                "about_title": "À propos",
                "about_version": "Version 2.0",
                "about_desc": "Une calculatrice spécialisée pour l'addition de nombres avec fractions.",
                "about_shortcuts": "Raccourcis clavier:",
                "shortcut_digits": "0-9 : Saisir les chiffres",
                "shortcut_decimal": ". ou , : Ajouter une virgule décimale",
                "shortcut_enter": "Entrée : Valider le nombre",
                "shortcut_insert": "Insert ou Ctrl+0 : Enregistrer le total",
                "shortcut_escape": "Échap : Effacer les nombres",
                "shortcut_delete": "Suppr : Effacer les totaux",
                "shortcut_backspace": "Retour arrière : Supprimer le dernier caractère",
                "shortcut_undo": "Ctrl+Z : Annuler (Undo)",
                "shortcut_redo": "Ctrl+Y : Rétablir (Redo)",
                "about_voice": "Commandes vocales:",
                "voice_numbers": "Chiffres : zéro, un, deux, trois...",
                "voice_decimal": "Virgule : 'virgule' ou 'point'",
                "voice_fractions": "Fractions : 'virgule vingt-cinq', 'virgule cinq', 'virgule soixante-quinze'",
                "voice_validate": "Valider : 'valider', 'entrer', 'ok'",
                "voice_delete": "Effacer : 'effacer', 'supprimer'",
                "voice_new": "Nouveau calcul : 'nouveau'",
                "voice_total": "Enregistrer total : 'total'",
                "save_dialog": "Enregistrer sous",
                "excel_files": "Fichiers Excel (*.xlsx)",
                "history_numbers_header": "Historique des Nombres",
                "history_totals_header": "Historique des Totaux",
                "mic_button": "🎤 Micro",
                "mic_on": "🎤 Écoute...",
                "mic_off": "🎤 Micro",
                "mic_tooltip_off": "Activer la dictée vocale",
                "mic_tooltip_on": "Désactiver la dictée vocale",
                "mic_error": "Erreur Microphone",
                "mic_not_available": "Vosk ou PyAudio n'est pas installé.\nInstallez avec: pip install vosk pyaudio",
                "mic_model_missing": "Modèle Vosk non trouvé.",
                "statistics": "Statistiques",
                "stats_title": "Statistiques des Totaux",
                "mean": "Moyenne",
                "std_dev": "Écart-type",
                "no_data": "Aucune donnée",
                "no_data_msg": "Aucun total enregistré pour afficher les statistiques.\nEnregistrez au moins un total avec 'Insert' ou 'Ctrl+0'.",
                "voice_model": "Modèle Vocal",
                "arabic_tunisia": "Arabe Tunisien (Linto)",
                "arabic_standard": "Arabe Standard (MGB2)",
                "french_model": "Français",
                "english_model": "Anglais",
                "model_changed": "Modèle changé",
                "model_changed_msg": "Le modèle vocal a été changé en : {0}\nRedémarrez le microphone pour appliquer les changements.",
                "model_error": "Erreur de modèle",
                "model_not_found": "Le modèle '{0}' n'a pas été trouvé.",
                "download_model": "Télécharger le modèle",
                "download_model_msg": "Le modèle '{0}' n'est pas installé.\nVoulez-vous le télécharger maintenant ?",
                "downloading": "Téléchargement...",
                "download_success": "Téléchargement réussi",
                "download_success_msg": "Le modèle a été téléchargé et installé avec succès !",
                "download_failed": "Échec du téléchargement",
                "cancel": "Annuler",
                # ── Confirm dialogs ──
                "confirm_clear_numbers_title": "Effacer les nombres",
                "confirm_clear_numbers_msg": "Voulez-vous effacer tous les nombres et réinitialiser le total ?",
                "confirm_clear_totals_title": "Effacer les totaux",
                "confirm_clear_totals_msg": "Voulez-vous effacer tout l'historique des totaux ?",
                "confirm_backspace_title": "Supprimer l'entrée",
                "confirm_backspace_row_msg": "Voulez-vous supprimer la ligne sélectionnée ?",
                # ── Undo / Redo ──
                "undo": "Annuler",
                "redo": "Rétablir",
                "nothing_to_undo": "Rien à annuler",
                "nothing_to_redo": "Rien à rétablir",
                # ── Help / Shortcuts dialog ──
                "shortcuts_title": "Raccourcis & Manipulations",
                "shortcuts_keyboard_section": "⌨️  Raccourcis clavier",
                "shortcuts_mouse_section": "🖱️  Manipulations à la souris",
                "mouse_left_digit": "Clic gauche sur un chiffre : ajouter le chiffre à la saisie",
                "mouse_right_digit": "Clic droit sur un chiffre : ajouter le chiffre ET valider immédiatement",
                "mouse_fraction": "Clic sur une fraction (0.25 / 0.5 / 0.75) : ajouter la fraction et valider",
                "mouse_total_btn": "Clic sur le bouton total : enregistrer le total courant dans l'historique",
                "mouse_table_dblclick": "Double-clic sur une valeur dans l'historique des nombres : modifier la valeur",
                "mouse_table_select_backspace": "Sélectionner une ligne + ⌫ : supprimer cette ligne",
            },
            "ar": {
                "title": "حاسبة النقاط",
                "file_menu": "ملف",
                "language_menu": "اللغة",
                "help_menu": "مساعدة",
                "shortcuts_action": "اختصارات & عمليات الفأرة",
                "export": "تصدير إلى Excel",
                "about": "حول",
                "quit": "خروج",
                "french": "Français",
                "arabic": "العربية",
                "english": "English",
                "clear_numbers": "مسح الأرقام",
                "clear_totals": "مسح المجاميع",
                "current_input": ":الإدخال الحالي",
                "counter_header": "#",
                "value_header": "القيمة",
                "total_header": "المجموع",
                "tooltip_history": "سجل الأرقام (انقر مرتين للتعديل)",
                "tooltip_totals": "سجل المجاميع المحفوظة",
                "tooltip_clear_numbers": "مسح قائمة الأرقام المضافة (Échap)",
                "tooltip_clear_totals": "مسح سجل المجاميع (Suppr)",
                "tooltip_total": "المجموع الحالي (Insert أو Ctrl+0 للحفظ)",
                "tooltip_enter": "تأكيد الرقم (Enter)",
                "tooltip_backspace": "مسح (Retour arrière)",
                "export_success": "نجح التصدير",
                "export_success_msg": "تم تصدير البيانات إلى:\n{0}",
                "export_error": "خطأ",
                "export_error_msg": "خطأ أثناء التصدير:\n{0}",
                "about_title": "حول",
                "about_version": "الإصدار 2.0",
                "about_desc": "آلة حاسبة متخصصة في جمع الأرقام مع الكسور.",
                "about_shortcuts": ":اختصارات لوحة المفاتيح",
                "shortcut_digits": "0-9 : إدخال الأرقام",
                "shortcut_decimal": ". أو , : إضافة فاصلة عشرية",
                "shortcut_enter": "Enter : تأكيد الرقم",
                "shortcut_insert": "Insert أو Ctrl+0 : حفظ المجموع",
                "shortcut_escape": "Échap : مسح الأرقام",
                "shortcut_delete": "Suppr : مسح المجاميع",
                "shortcut_backspace": "Retour arrière : حذف آخر حرف",
                "shortcut_undo": "Ctrl+Z : تراجع (Undo)",
                "shortcut_redo": "Ctrl+Y : إعادة (Redo)",
                "about_voice": ":الأوامر الصوتية",
                "voice_numbers": "الأرقام : صفر، واحد، اثنان، ثلاثة...",
                "voice_decimal": "الفاصلة : 'فاصل' أو 'فاصلة'",
                "voice_fractions": "الكسور : 'فاصل ربع'، 'فاصل نصف'، 'فاصل ثلاثة أرباع'",
                "voice_validate": "التأكيد : 'تأكيد', 'ok'",
                "voice_delete": "المسح : 'مسح'",
                "voice_new": "حساب جديد : 'جديد'",
                "voice_total": "حفظ المجموع : 'مجموع', 'المجموع'",
                "save_dialog": "حفظ باسم",
                "excel_files": "(*.xlsx) ملفات Excel",
                "history_numbers_header": "سجل الأرقام",
                "history_totals_header": "سجل المجاميع",
                "mic_button": "🎤 ميكروفون",
                "mic_on": "🎤 استماع...",
                "mic_off": "🎤 ميكروفون",
                "mic_tooltip_off": "تفعيل الإملاء الصوتي",
                "mic_tooltip_on": "إيقاف الإملاء الصوتي",
                "mic_error": "خطأ في الميكروفون",
                "mic_not_available": "Vosk أو PyAudio غير مثبت.\nقم بالتثبيت باستخدام: pip install vosk pyaudio",
                "mic_model_missing": "نموذج Vosk غير موجود.",
                "statistics": "إحصائيات",
                "stats_title": "إحصائيات المجاميع",
                "mean": "المتوسط",
                "std_dev": "الانحراف المعياري",
                "no_data": "لا توجد بيانات",
                "no_data_msg": "لا توجد مجاميع محفوظة لعرض الإحصائيات.\nاحفظ مجموعًا واحدًا على الأقل باستخدام 'Insert' أو 'Ctrl+0'.",
                "voice_model": "نموذج الصوت",
                "arabic_tunisia": "تونسي (Linto)",
                "arabic_standard": "فصحى (MGB2)",
                "french_model": "فرنسي",
                "english_model": "إنجليزي",
                "model_changed": "تم تغيير النموذج",
                "model_changed_msg": "تم تغيير النموذج الصوتي إلى: {0}\nأعد تشغيل الميكروفون لتطبيق التغييرات.",
                "model_error": "خطأ في النموذج",
                "model_not_found": "النموذج '{0}' غير موجود.",
                "download_model": "تنزيل النموذج",
                "download_model_msg": "النموذج '{0}' غير مثبت.\nهل تريد تنزيله الآن؟",
                "downloading": "...جاري التنزيل",
                "download_success": "نجح التنزيل",
                "download_success_msg": "!تم تنزيل النموذج وتثبيته بنجاح",
                "download_failed": "فشل التنزيل",
                "cancel": "إلغاء",
                "confirm_clear_numbers_title": "مسح الأرقام",
                "confirm_clear_numbers_msg": "هل تريد مسح جميع الأرقام وإعادة تعيين المجموع؟",
                "confirm_clear_totals_title": "مسح المجاميع",
                "confirm_clear_totals_msg": "هل تريد مسح سجل المجاميع بالكامل؟",
                "confirm_backspace_title": "حذف الإدخال",
                "confirm_backspace_row_msg": "هل تريد حذف الصف المحدد؟",
                "undo": "تراجع",
                "redo": "إعادة",
                "nothing_to_undo": "لا يوجد شيء للتراجع عنه",
                "nothing_to_redo": "لا يوجد شيء لإعادته",
                "shortcuts_title": "اختصارات & عمليات الفأرة",
                "shortcuts_keyboard_section": "⌨️  اختصارات لوحة المفاتيح",
                "shortcuts_mouse_section": "🖱️  عمليات الفأرة",
                "mouse_left_digit": "نقرة يسار على رقم : إضافة الرقم للإدخال",
                "mouse_right_digit": "نقرة يمين على رقم : إضافة الرقم والتأكيد فورًا",
                "mouse_fraction": "نقر على كسر (0.25/0.5/0.75) : إضافة الكسر والتأكيد",
                "mouse_total_btn": "نقر على زر المجموع : حفظ المجموع الحالي في السجل",
                "mouse_table_dblclick": "نقر مزدوج على قيمة في سجل الأرقام : تعديل القيمة",
                "mouse_table_select_backspace": "تحديد صف + ⌫ : حذف هذا الصف",
            },
            "en": {
                "title": "Calculateur de score",
                "file_menu": "File",
                "language_menu": "Language",
                "help_menu": "Help",
                "shortcuts_action": "Shortcuts & Mouse Actions",
                "export": "Export to Excel",
                "about": "About",
                "quit": "Quit",
                "french": "Français",
                "arabic": "العربية",
                "english": "English",
                "clear_numbers": "Clear Numbers",
                "clear_totals": "Clear Totals",
                "current_input": "Current input:",
                "counter_header": "#",
                "value_header": "Value",
                "total_header": "Total",
                "tooltip_history": "Number history (double-click to edit)",
                "tooltip_totals": "Saved totals history",
                "tooltip_clear_numbers": "Clear the list of added numbers (Escape)",
                "tooltip_clear_totals": "Clear totals history (Delete)",
                "tooltip_total": "Current total (Insert or Ctrl+0 to save)",
                "tooltip_enter": "Validate number (Enter)",
                "tooltip_backspace": "Delete (Backspace)",
                "export_success": "Export successful",
                "export_success_msg": "Data has been exported to:\n{0}",
                "export_error": "Error",
                "export_error_msg": "Error during export:\n{0}",
                "about_title": "About",
                "about_version": "Version 2.0",
                "about_desc": "A specialized calculator for adding numbers with fractions.",
                "about_shortcuts": "Keyboard shortcuts:",
                "shortcut_digits": "0-9: Enter digits",
                "shortcut_decimal": ". or ,: Add decimal point",
                "shortcut_enter": "Enter: Validate number",
                "shortcut_insert": "Insert or Ctrl+0: Save total",
                "shortcut_escape": "Escape: Clear numbers",
                "shortcut_delete": "Delete: Clear totals",
                "shortcut_backspace": "Backspace: Delete last character",
                "shortcut_undo": "Ctrl+Z: Undo",
                "shortcut_redo": "Ctrl+Y: Redo",
                "about_voice": "Voice commands:",
                "voice_numbers": "Numbers: zero, one, two, three...",
                "voice_decimal": "Decimal: 'point'",
                "voice_fractions": "Fractions: 'point twenty-five', 'point five', 'point seventy-five'",
                "voice_validate": "Validate: 'enter', 'ok'",
                "voice_delete": "Delete: 'delete', 'clear'",
                "voice_new": "New calculation: 'new'",
                "voice_total": "Save total: 'total'",
                "save_dialog": "Save as",
                "excel_files": "Excel Files (*.xlsx)",
                "history_numbers_header": "Number History",
                "history_totals_header": "Totals History",
                "mic_button": "🎤 Mic",
                "mic_on": "🎤 Listening...",
                "mic_off": "🎤 Mic",
                "mic_tooltip_off": "Enable voice dictation",
                "mic_tooltip_on": "Disable voice dictation",
                "mic_error": "Microphone Error",
                "mic_not_available": "Vosk or PyAudio is not installed.\nInstall with: pip install vosk pyaudio",
                "mic_model_missing": "Vosk model not found.",
                "statistics": "Statistics",
                "stats_title": "Totals Statistics",
                "mean": "Mean",
                "std_dev": "Standard Deviation",
                "no_data": "No Data",
                "no_data_msg": "No totals saved to display statistics.\nSave at least one total with 'Insert' or 'Ctrl+0'.",
                "voice_model": "Voice Model",
                "arabic_tunisia": "Tunisian (Linto)",
                "arabic_standard": "Standard (MGB2)",
                "french_model": "French",
                "english_model": "English",
                "model_changed": "Model Changed",
                "model_changed_msg": "Voice model has been changed to: {0}\nRestart the microphone to apply changes.",
                "model_error": "Model Error",
                "model_not_found": "Model '{0}' was not found.",
                "download_model": "Download Model",
                "download_model_msg": "Model '{0}' is not installed.\nDo you want to download it now?",
                "downloading": "Downloading...",
                "download_success": "Download Successful",
                "download_success_msg": "The model has been downloaded and installed successfully!",
                "download_failed": "Download Failed",
                "cancel": "Cancel",
                "confirm_clear_numbers_title": "Clear Numbers",
                "confirm_clear_numbers_msg": "Do you want to clear all numbers and reset the total?",
                "confirm_clear_totals_title": "Clear Totals",
                "confirm_clear_totals_msg": "Do you want to clear the entire totals history?",
                "confirm_backspace_title": "Delete Entry",
                "confirm_backspace_row_msg": "Do you want to delete the selected row?",
                "undo": "Undo",
                "redo": "Redo",
                "nothing_to_undo": "Nothing to undo",
                "nothing_to_redo": "Nothing to redo",
                "shortcuts_title": "Shortcuts & Mouse Actions",
                "shortcuts_keyboard_section": "⌨️  Keyboard Shortcuts",
                "shortcuts_mouse_section": "🖱️  Mouse Actions",
                "mouse_left_digit": "Left-click on a digit: add digit to current input",
                "mouse_right_digit": "Right-click on a digit: add digit AND validate immediately",
                "mouse_fraction": "Click on a fraction button (0.25/0.5/0.75): add fraction and validate",
                "mouse_total_btn": "Click on the total button: save current total to history",
                "mouse_table_dblclick": "Double-click on a value in the number history: edit the value",
                "mouse_table_select_backspace": "Select a row + ⌫: delete that row",
            }
        }
        self.initUI()

    def tr(self, key):
        return self.translations[self.current_language].get(key, key)

    # ── Undo / Redo helpers ──────────────────────────────────────────────────

    def _snapshot(self):
        """Return a deep copy of the current state."""
        return {
            'historique': list(self.historique),
            'total': self.total,
            'history_of_totals': list(self.history_of_totals),
            'current_number': self.current_number,
        }

    def _push_undo(self):
        """Save current state to undo stack and clear redo stack."""
        self._undo_stack.append(self._snapshot())
        self._redo_stack.clear()
        self._refresh_undo_redo_actions()

    def _restore(self, state):
        self.historique = list(state['historique'])
        self.total = state['total']
        self.history_of_totals = list(state['history_of_totals'])
        self.current_number = state['current_number']
        self.update_display()
        self.update_current_display()

    def undo(self):
        if not self._undo_stack:
            QMessageBox.information(self, self.tr("undo"), self.tr("nothing_to_undo"))
            return
        self._redo_stack.append(self._snapshot())
        state = self._undo_stack.pop()
        self._restore(state)
        self._refresh_undo_redo_actions()

    def redo(self):
        if not self._redo_stack:
            QMessageBox.information(self, self.tr("redo"), self.tr("nothing_to_redo"))
            return
        self._undo_stack.append(self._snapshot())
        state = self._redo_stack.pop()
        self._restore(state)
        self._refresh_undo_redo_actions()

    def _refresh_undo_redo_actions(self):
        self.undo_action.setEnabled(bool(self._undo_stack))
        self.redo_action.setEnabled(bool(self._redo_stack))

    # ────────────────────────────────────────────────────────────────────────

    def initUI(self):
        self.setWindowTitle(self.tr("title"))
        self.setStyleSheet("""
            QWidget {
                background-color: #f5f5f5;
                font-family: 'Segoe UI', Arial, sans-serif;
            }
            QMenuBar::item {
                background-color: transparent;
                padding: 8px 12px;
                color: black;
            }
            QMenuBar::item:selected {
                background-color: #3498db;
                border-radius: 4px;
                color: white;
            }
            QMenu {
                background-color: white;
                border: 1px solid #bdc3c7;
                padding: 5px;
            }
            QMenu::item {
                padding: 8px 25px;
                color: #2c3e50;
            }
            QMenu::item:selected {
                background-color: #3498db;
                color: white;
                border-radius: 4px;
            }
            QMenu::item:disabled {
                color: #aaa;
            }
            QTableWidget {
                background-color: white;
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                gridline-color: #ecf0f1;
                font-size: 12px;
            }
            QTableWidget::item { padding: 5px; }
            QTableWidget::item:selected { background-color: #3498db; color: white; }
            QHeaderView::section {
                background-color: #3498db;
                color: white;
                padding: 8px;
                border: none;
                font-weight: bold;
                font-size: 11px;
            }
            QPushButton {
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                padding: 5px;
                background-color: white;
                font-weight: bold;
            }
            QPushButton:hover { background-color: #ecf0f1; border-color: #95a5a6; }
            QPushButton:pressed { background-color: #bdc3c7; }
            QLabel { color: #2c3e50; }
        """)

        screen = QDesktopWidget().screenGeometry()
        self.setGeometry(screen.width() // 3, 40, 700, screen.height())
        self.setFixedHeight(9 * screen.height() // 10)

        main_layout = QVBoxLayout()

        # ── Menu bar ─────────────────────────────────────────────────────────
        self.menubar = QMenuBar(self)

        # File menu
        self.file_menu = self.menubar.addMenu(self.tr("file_menu"))
        self.export_action = QAction(self.tr("export"), self)
        self.export_action.triggered.connect(self.export_to_excel)
        self.file_menu.addAction(self.export_action)
        self.stats_action = QAction(self.tr("statistics"), self)
        self.stats_action.triggered.connect(self.show_statistics)
        self.file_menu.addAction(self.stats_action)
        self.file_menu.addSeparator()

        # Undo / Redo inside File menu
        self.undo_action = QAction(self.tr("undo") + "    Ctrl+Z", self)
        self.undo_action.setEnabled(False)
        self.undo_action.triggered.connect(self.undo)
        self.file_menu.addAction(self.undo_action)

        self.redo_action = QAction(self.tr("redo") + "    Ctrl+Y", self)
        self.redo_action.setEnabled(False)
        self.redo_action.triggered.connect(self.redo)
        self.file_menu.addAction(self.redo_action)
        self.file_menu.addSeparator()

        self.quit_action = QAction(self.tr("quit"), self)
        self.quit_action.triggered.connect(self.close)
        self.file_menu.addAction(self.quit_action)

        # Language menu
        self.language_menu = self.menubar.addMenu(self.tr("language_menu"))
        self.language_group = QActionGroup(self)
        self.language_group.setExclusive(True)

        self.french_action = QAction(self.tr("french"), self)
        self.french_action.setCheckable(True)
        self.french_action.setChecked(True)
        self.french_action.triggered.connect(lambda: self.change_language("fr"))
        self.language_group.addAction(self.french_action)
        self.language_menu.addAction(self.french_action)

        self.language_menu.addSeparator()
        self.arabic_action = QAction(self.tr("arabic"), self)
        self.arabic_action.setCheckable(True)
        self.arabic_action.triggered.connect(lambda: self.change_language("ar"))
        self.language_group.addAction(self.arabic_action)
        self.language_menu.addAction(self.arabic_action)

        self.arabic_tunisia_action = QAction("    → " + self.tr("arabic_tunisia"), self)
        self.arabic_tunisia_action.triggered.connect(lambda: self.change_voice_model("vosk-model-small-ar-tn-0.1-linto"))
        self.language_menu.addAction(self.arabic_tunisia_action)

        self.arabic_standard_action = QAction("    → " + self.tr("arabic_standard"), self)
        self.arabic_standard_action.triggered.connect(lambda: self.change_voice_model("vosk-model-ar-mgb2-0.4"))
        self.language_menu.addAction(self.arabic_standard_action)

        self.language_menu.addSeparator()
        self.english_action = QAction(self.tr("english"), self)
        self.english_action.setCheckable(True)
        self.english_action.triggered.connect(lambda: self.change_language("en"))
        self.language_group.addAction(self.english_action)
        self.language_menu.addAction(self.english_action)

        # ── Help menu (NEW) ──────────────────────────────────────────────────
        self.help_menu = self.menubar.addMenu(self.tr("help_menu"))

        self.shortcuts_action = QAction(self.tr("shortcuts_action"), self)
        self.shortcuts_action.triggered.connect(self.show_shortcuts)
        self.help_menu.addAction(self.shortcuts_action)

        self.help_menu.addSeparator()

        self.about_action = QAction(self.tr("about"), self)
        self.about_action.triggered.connect(self.show_about)
        self.help_menu.addAction(self.about_action)

        main_layout.setMenuBar(self.menubar)

        layout = QVBoxLayout()

        # ── History tables ───────────────────────────────────────────────────
        historique_layout = QHBoxLayout()

        self.historique_table = QTableWidget(self)
        self.historique_table.setColumnCount(2)
        self.historique_table.setHorizontalHeaderLabels([self.tr("counter_header"), self.tr("value_header")])
        self.historique_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.historique_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.historique_table.setFixedWidth(300)
        self.historique_table.setToolTip(self.tr("tooltip_history"))
        self.historique_table.itemChanged.connect(self.on_historique_changed)

        self.history_table = QTableWidget(self)
        self.history_table.setColumnCount(2)
        self.history_table.setHorizontalHeaderLabels([self.tr("total_header"), self.tr("value_header")])
        self.history_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.history_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.history_table.setFixedWidth(300)
        self.history_table.setToolTip(self.tr("tooltip_totals"))

        historique_layout.addWidget(self.historique_table)
        historique_layout.addWidget(self.history_table)
        layout.addLayout(historique_layout)

        # ── Current input row ────────────────────────────────────────────────
        current_layout = QHBoxLayout()
        self.current_label = QLabel(self.tr("current_input"), self)
        self.current_label.setFont(QFont("Arial", 11, QFont.Bold))
        self.current_label.setStyleSheet("color: #34495e;")

        self.current_input_label = QPushButton("", self)
        self.current_input_label.setFixedHeight(40)
        self.current_input_label.setFont(QFont("Arial", 18, QFont.Bold))
        self.current_input_label.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                           stop:0 #fff9c4, stop:1 #fff176);
                color: #2c3e50; font-size: 18px; font-weight: bold;
                padding: 5px 10px; border: 2px solid #fbc02d; border-radius: 8px;
                text-align: right;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                           stop:0 #ffe082, stop:1 #ffd54f);
                border-color: #f9a825;
            }
            QPushButton:pressed {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                           stop:0 #ffd54f, stop:1 #ffca28);
            }
        """)
        self.current_input_label.setToolTip(self.tr("tooltip_enter"))
        self.current_input_label.clicked.connect(self.validate_number)

        self.mic_button = QPushButton(self.tr("mic_off"), self)
        self.mic_button.setFixedHeight(40)
        self.mic_button.setFixedWidth(130)
        self.mic_button.setFont(QFont("Arial", 11, QFont.Bold))
        self.mic_button.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                           stop:0 #64b5f6, stop:1 #2196f3);
                color: white; border: 2px solid #1976d2; border-radius: 8px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                           stop:0 #42a5f5, stop:1 #1976d2);
            }
        """)
        self.mic_button.setToolTip(self.tr("mic_tooltip_off"))
        self.mic_button.clicked.connect(self.toggle_microphone)

        current_layout.addWidget(self.current_label)
        current_layout.addWidget(self.current_input_label)
        current_layout.addWidget(self.mic_button)
        layout.addLayout(current_layout)

        # ── Total + Clear buttons ────────────────────────────────────────────
        total_layout = QHBoxLayout()

        self.clear_numbers_button = QPushButton(self.tr("clear_numbers"), self)
        self.clear_numbers_button.clicked.connect(self.clear_numbers)
        self.clear_numbers_button.setFixedHeight(50)
        self.clear_numbers_button.setFont(QFont("Arial", 12, QFont.Bold))
        self.clear_numbers_button.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                           stop:0 #81c784, stop:1 #66bb6a);
                color: white; border: 2px solid #4caf50; border-radius: 10px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                           stop:0 #66bb6a, stop:1 #4caf50);
            }
        """)
        self.clear_numbers_button.setToolTip(self.tr("tooltip_clear_numbers"))

        self.total_value_button = QPushButton(f'{self.total:.2f}', self)
        self.total_value_button.setFixedHeight(60)
        self.total_value_button.setFont(QFont("Arial", 24, QFont.Bold))
        self.total_value_button.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                           stop:0 #424242, stop:1 #212121);
                color: #4CAF50; border: 3px solid #4CAF50; border-radius: 12px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                           stop:0 #616161, stop:1 #424242);
                color: #66BB6A; border-color: #66BB6A;
            }
            QPushButton:pressed {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                           stop:0 #212121, stop:1 #000000);
            }
        """)
        self.total_value_button.setToolTip(self.tr("tooltip_total"))
        self.total_value_button.clicked.connect(self.save_total_from_button)

        self.clear_totals_button = QPushButton(self.tr("clear_totals"), self)
        self.clear_totals_button.clicked.connect(self.clear_totals)
        self.clear_totals_button.setFixedHeight(50)
        self.clear_totals_button.setFont(QFont("Arial", 12, QFont.Bold))
        self.clear_totals_button.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                           stop:0 #90a4ae, stop:1 #78909c);
                color: white; border: 2px solid #607d8b; border-radius: 10px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                           stop:0 #78909c, stop:1 #607d8b);
            }
        """)
        self.clear_totals_button.setToolTip(self.tr("tooltip_clear_totals"))

        total_layout.addWidget(self.clear_numbers_button)
        total_layout.addWidget(self.total_value_button)
        total_layout.addWidget(self.clear_totals_button)
        layout.addLayout(total_layout)

        # ── Digit / Fraction grid ────────────────────────────────────────────
        boutons_layout = QGridLayout()
        left_buttons = [
            (7, 0, 0), (8, 0, 1), (9, 0, 2),
            (4, 1, 0), (5, 1, 1), (6, 1, 2),
            (1, 2, 0), (2, 2, 1), (3, 2, 2),
            (0, 3, 0)
        ]
        for valeur, row, col in left_buttons:
            button = QPushButton(str(valeur), self)
            button.setFixedHeight(75)
            button.setFixedWidth(105)
            button.setFont(QFont("Arial", 18, QFont.Bold))
            button.setStyleSheet("""
                QPushButton {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                               stop:0 #ffffff, stop:1 #e3f2fd);
                    color: #1976d2; border: 3px solid #64b5f6; border-radius: 12px;
                }
                QPushButton:hover {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                               stop:0 #e3f2fd, stop:1 #bbdefb);
                    border-color: #42a5f5;
                }
                QPushButton:pressed { background: #90caf9; }
            """)
            button.clicked.connect(lambda _, v=valeur: self.add_digit(v))
            button.setContextMenuPolicy(Qt.CustomContextMenu)
            button.customContextMenuRequested.connect(lambda pos, v=valeur: self.add_and_validate_digit(v))
            boutons_layout.addWidget(button, row, col)

        self.decimal_button = QPushButton(".", self)
        self.decimal_button.setFixedHeight(75)
        self.decimal_button.setFixedWidth(105)
        self.decimal_button.setFont(QFont("Arial", 24, QFont.Bold))
        self.decimal_button.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                           stop:0 #c8e6c9, stop:1 #a5d6a7);
                color: #2e7d32; border: 3px solid #66bb6a; border-radius: 12px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                           stop:0 #a5d6a7, stop:1 #81c784);
            }
            QPushButton:pressed { background: #66bb6a; }
        """)
        self.decimal_button.clicked.connect(self.add_decimal_point)
        boutons_layout.addWidget(self.decimal_button, 3, 1)

        fractions = [0.75, 0.5, 0.25]
        for i, frac in enumerate(fractions):
            button = QPushButton(str(frac), self)
            button.setFixedHeight(75)
            button.setFixedWidth(105)
            button.setFont(QFont("Arial", 16, QFont.Bold))
            button.setStyleSheet("""
                QPushButton {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                               stop:0 #e1bee7, stop:1 #ce93d8);
                    color: #6a1b9a; border: 3px solid #ab47bc; border-radius: 12px;
                }
                QPushButton:hover {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                               stop:0 #ce93d8, stop:1 #ba68c8);
                }
                QPushButton:pressed { background: #ab47bc; }
            """)
            button.clicked.connect(lambda _, f=frac: self.add_fraction(f))
            boutons_layout.addWidget(button, i, 4)

        self.enter_button = QPushButton("✓", self)
        self.enter_button.setFixedHeight(75)
        self.enter_button.setFixedWidth(105)
        self.enter_button.setFont(QFont("Arial", 24, QFont.Bold))
        self.enter_button.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                           stop:0 #81c784, stop:1 #66bb6a);
                color: white; border: 3px solid #4caf50; border-radius: 12px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                           stop:0 #66bb6a, stop:1 #4caf50);
            }
            QPushButton:pressed { background: #388e3c; }
        """)
        self.enter_button.setToolTip(self.tr("tooltip_enter"))
        self.enter_button.clicked.connect(self.validate_number)
        boutons_layout.addWidget(self.enter_button, 3, 2)

        self.backspace_button = QPushButton("⌫", self)
        self.backspace_button.setFixedHeight(75)
        self.backspace_button.setFixedWidth(105)
        self.backspace_button.setFont(QFont("Arial", 20, QFont.Bold))
        self.backspace_button.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                           stop:0 #ef5350, stop:1 #e53935);
                color: white; border: 3px solid #c62828; border-radius: 12px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                           stop:0 #e53935, stop:1 #c62828);
            }
            QPushButton:pressed { background: #b71c1c; }
        """)
        self.backspace_button.setToolTip(self.tr("tooltip_backspace"))
        self.backspace_button.clicked.connect(self.backspace)
        boutons_layout.addWidget(self.backspace_button, 3, 4)

        layout.addLayout(boutons_layout)
        main_layout.addLayout(layout)
        self.setLayout(main_layout)
        self.setFocusPolicy(Qt.StrongFocus)

        self.init_vosk()

        self.audio_timer = QTimer()
        self.audio_timer.timeout.connect(self.process_audio_queue)
        self.audio_timer.start(100)

    # ── Help: Shortcuts dialog ───────────────────────────────────────────────

    def show_shortcuts(self):
        """Display a dialog explaining all keyboard shortcuts and mouse actions."""
        dialog = QDialog(self)
        dialog.setWindowTitle(self.tr("shortcuts_title"))
        dialog.setMinimumWidth(560)
        dialog.setStyleSheet("""
            QDialog { background-color: #f8f9fa; }
            QLabel { color: #2c3e50; }
            QScrollArea { border: none; background-color: transparent; }
            QScrollBar:vertical {
                border: none;
                background: #f0f0f0;
                width: 12px;
                margin: 0px;
            }
            QScrollBar::handle:vertical {
                background: #bdc3c7;
                border-radius: 6px;
                min-height: 30px;
            }
            QScrollBar::handle:vertical:hover { background: #95a5a6; }
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
                border: none;
                background: none;
            }
        """)

        # Main layout for dialog (contains scroll area + fixed close button)
        main_layout = QVBoxLayout()
        main_layout.setContentsMargins(20, 20, 20, 20)
        main_layout.setSpacing(15)

        # Scroll area
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)

        # Container widget for all scrollable content
        container = QWidget()
        layout = QVBoxLayout(container)
        layout.setSpacing(12)
        layout.setContentsMargins(0, 0, 0, 0)

        def section_label(text):
            lbl = QLabel(text)
            lbl.setFont(QFont("Arial", 12, QFont.Bold))
            lbl.setStyleSheet("""
                background-color: #3498db; color: white;
                padding: 6px 10px; border-radius: 6px;
            """)
            return lbl

        def row_label(icon, text):
            lbl = QLabel(f"  {icon}  {text}")
            lbl.setFont(QFont("Arial", 10))
            lbl.setStyleSheet("padding: 3px 6px;")
            lbl.setWordWrap(True)
            return lbl

        # Keyboard section
        layout.addWidget(section_label(self.tr("shortcuts_keyboard_section")))
        shortcuts = [
            ("🔢", "0 – 9", self.tr("shortcut_digits")),
            ("🔸", ". / ,", self.tr("shortcut_decimal")),
            ("↵", "Enter", self.tr("shortcut_enter")),
            ("💾", "Insert / Ctrl+0", self.tr("shortcut_insert")),
            ("🗑️", "Escape", self.tr("shortcut_escape")),
            ("🗑️", "Delete", self.tr("shortcut_delete")),
            ("⌫", "Backspace", self.tr("shortcut_backspace")),
            ("↩️", "Ctrl+Z", self.tr("shortcut_undo")),
            ("↪️", "Ctrl+Y", self.tr("shortcut_redo")),
        ]
        for icon, key, desc in shortcuts:
            row = QHBoxLayout()
            key_lbl = QLabel(key)
            key_lbl.setFont(QFont("Courier New", 10, QFont.Bold))
            key_lbl.setFixedWidth(130)
            key_lbl.setStyleSheet("""
                background-color: #ecf0f1; border: 1px solid #bdc3c7;
                border-radius: 4px; padding: 3px 8px; color: #2c3e50;
            """)
            key_lbl.setAlignment(Qt.AlignCenter)
            desc_lbl = QLabel(f"{icon}  {desc}")
            desc_lbl.setFont(QFont("Arial", 10))
            desc_lbl.setWordWrap(True)
            row.addWidget(key_lbl)
            row.addWidget(desc_lbl, 1)
            layout.addLayout(row)

        # Mouse section
        layout.addSpacing(8)
        layout.addWidget(section_label(self.tr("shortcuts_mouse_section")))
        mouse_items = [
            ("🖱️ LClick", self.tr("mouse_left_digit")),
            ("🖱️ RClick", self.tr("mouse_right_digit")),
            ("🔵", self.tr("mouse_fraction")),
            ("🟢", self.tr("mouse_total_btn")),
            ("✏️", self.tr("mouse_table_dblclick")),
            ("🗑️", self.tr("mouse_table_select_backspace")),
        ]
        for icon, desc in mouse_items:
            row = QHBoxLayout()
            icon_lbl = QLabel(icon)
            icon_lbl.setFont(QFont("Arial", 10, QFont.Bold))
            icon_lbl.setFixedWidth(80)
            icon_lbl.setAlignment(Qt.AlignCenter)
            icon_lbl.setStyleSheet("""
                background-color: #ecf0f1; border: 1px solid #bdc3c7;
                border-radius: 4px; padding: 3px 6px; color: #2c3e50;
            """)
            desc_lbl = QLabel(desc)
            desc_lbl.setFont(QFont("Arial", 10))
            desc_lbl.setWordWrap(True)
            row.addWidget(icon_lbl)
            row.addWidget(desc_lbl, 1)
            layout.addLayout(row)

        # Add stretch to push content to the top if needed
        layout.addStretch()

        # Set container as scroll area's widget
        scroll.setWidget(container)

        # Add scroll area to main layout
        main_layout.addWidget(scroll, 1)

        # Close button (fixed at bottom)
        close_btn = QPushButton("✕  " + self.tr("cancel"), dialog)
        close_btn.setFixedHeight(40)
        close_btn.setFont(QFont("Arial", 11, QFont.Bold))
        close_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0,y1:0,x2:0,y2:1,
                                           stop:0 #3498db, stop:1 #2980b9);
                color: white; border: none; border-radius: 8px;
            }
            QPushButton:hover { background: #2980b9; }
        """)
        close_btn.clicked.connect(dialog.close)
        main_layout.addWidget(close_btn)

        dialog.setLayout(main_layout)
        dialog.exec_()
    
    # ── Vosk helpers (unchanged) ─────────────────────────────────────────────

    def init_vosk(self):
        if not VOSK_AVAILABLE:
            return
        try:
            default_model = "vosk-model-small-ar-tn-0.1-linto"
            appdata_path = Path(os.getenv('APPDATA')) / 'Calculator' / 'vosk-models'
            model_paths = [
                appdata_path / default_model,
                appdata_path / "vosk-model-ar-mgb2-0.4",
                appdata_path / "vosk-model-small-fr-0.22",
                Path(default_model),
                Path("vosk-model-ar-mgb2-0.4"),
                Path("model")
            ]
            for path in model_paths:
                if path.exists():
                    self.vosk_model = vosk.Model(str(path))
                    self.current_model_path = str(path)
                    return
        except Exception as e:
            print(f"Erreur lors du chargement du modèle Vosk: {e}")

    def get_model_path(self, model_name):
        appdata_path = Path(os.getenv('APPDATA')) / 'Calculator' / 'vosk-models' / model_name
        if appdata_path.exists():
            return str(appdata_path)
        local_path = Path(model_name)
        if local_path.exists():
            return str(local_path)
        return None

    def download_model(self, model_name):
        if model_name not in self.model_urls:
            QMessageBox.warning(self, self.tr("model_error"),
                                f"URL non disponible pour le modèle {model_name}")
            return
        self.progress_dialog = QProgressDialog(
            self.tr("downloading"), self.tr("cancel"), 0, 100, self)
        self.progress_dialog.setWindowTitle(self.tr("download_model"))
        self.progress_dialog.setWindowModality(Qt.WindowModal)
        self.progress_dialog.setMinimumDuration(0)
        self.progress_dialog.setValue(0)
        self.download_thread = ModelDownloader(model_name, self.model_urls[model_name])
        self.download_thread.progress.connect(self.progress_dialog.setValue)
        self.download_thread.status.connect(self.progress_dialog.setLabelText)
        self.download_thread.finished.connect(self.on_download_finished)
        self.progress_dialog.canceled.connect(self.download_thread.cancel)
        self.download_thread.start()

    def on_download_finished(self, success, message):
        self.progress_dialog.close()
        if success:
            QMessageBox.information(self, self.tr("download_success"),
                                    self.tr("download_success_msg"))
            model_path = message
            try:
                self.vosk_model = vosk.Model(model_path)
                self.current_model_path = model_path
                if self.pending_language_change:
                    lang_code = self.pending_language_change
                    self.pending_language_change = None
                    self.current_language = lang_code
                    if lang_code == "fr":
                        self.french_action.setChecked(True)
                    elif lang_code == "ar":
                        self.arabic_action.setChecked(True)
                    elif lang_code == "en":
                        self.english_action.setChecked(True)
                    self._refresh_all_ui_texts()
                    if lang_code == "ar":
                        self.setLayoutDirection(Qt.RightToLeft)
                    else:
                        self.setLayoutDirection(Qt.LeftToRight)
                    self.update_display()
            except Exception as e:
                QMessageBox.critical(self, self.tr("model_error"),
                                     f"Erreur lors du chargement du modèle:\n{str(e)}")
        else:
            QMessageBox.critical(self, self.tr("download_failed"), message)

    def change_voice_model(self, model_name):
        was_listening = self.is_listening
        if was_listening:
            self.stop_listening()
        model_path = self.get_model_path(model_name)
        if model_path is None:
            reply = QMessageBox.question(
                self, self.tr("download_model"),
                self.tr("download_model_msg").format(model_name),
                QMessageBox.Yes | QMessageBox.No)
            if reply == QMessageBox.Yes:
                self.download_model(model_name)
            return
        try:
            self.vosk_model = vosk.Model(model_path)
            self.current_model_path = model_path
            QMessageBox.information(self, self.tr("model_changed"),
                                    self.tr("model_changed_msg").format(model_name))
        except Exception as e:
            QMessageBox.critical(self, self.tr("model_error"),
                                 f"Erreur lors du changement de modèle:\n{str(e)}")

    def toggle_microphone(self):
        if not VOSK_AVAILABLE:
            QMessageBox.warning(self, self.tr("mic_error"), self.tr("mic_not_available"))
            return
        required_model = self.language_models.get(self.current_language)
        if required_model:
            model_path = self.get_model_path(required_model)
            if self.vosk_model is None or (self.current_model_path and required_model not in self.current_model_path):
                if model_path is None:
                    reply = QMessageBox.question(
                        self, self.tr("download_model"),
                        self.tr("download_model_msg").format(required_model),
                        QMessageBox.Yes | QMessageBox.No)
                    if reply == QMessageBox.Yes:
                        self.download_model(required_model)
                    return
                else:
                    try:
                        self.vosk_model = vosk.Model(model_path)
                        self.current_model_path = model_path
                    except Exception as e:
                        QMessageBox.critical(self, self.tr("model_error"),
                                             f"Erreur lors du chargement du modèle:\n{str(e)}")
                        return
        if self.vosk_model is None:
            default_model = self.language_models.get(self.current_language, "vosk-model-small-ar-tn-0.1-linto")
            reply = QMessageBox.question(
                self, self.tr("download_model"),
                self.tr("download_model_msg").format(default_model),
                QMessageBox.Yes | QMessageBox.No)
            if reply == QMessageBox.Yes:
                self.download_model(default_model)
            return
        if not self.is_listening:
            self.start_listening()
        else:
            self.stop_listening()

    def start_listening(self):
        try:
            p = pyaudio.PyAudio()
            has_input = False
            try:
                p.get_default_input_device_info()
                has_input = True
            except:
                for i in range(p.get_device_count()):
                    if p.get_device_info_by_index(i)['maxInputChannels'] > 0:
                        has_input = True
                        break
            p.terminate()
            if not has_input:
                QMessageBox.warning(self, self.tr("mic_error"),
                                    "Aucun microphone détecté.")
                return
            self.is_listening = True
            self.mic_button.setText(self.tr("mic_on"))
            self.mic_button.setStyleSheet("""
                QPushButton {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                               stop:0 #ef5350, stop:1 #e53935);
                    color: white; border: 2px solid #c62828; border-radius: 8px;
                }
                QPushButton:hover {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                               stop:0 #e53935, stop:1 #c62828);
                }
            """)
            self.mic_button.setToolTip(self.tr("mic_tooltip_on"))
            self.audio_thread = threading.Thread(target=self.audio_listener_thread, daemon=True)
            self.audio_thread.start()
        except Exception as e:
            self.is_listening = False
            QMessageBox.critical(self, self.tr("mic_error"), f"Erreur: {str(e)}")

    def stop_listening(self):
        self.is_listening = False
        self.mic_button.setText(self.tr("mic_off"))
        self.mic_button.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                           stop:0 #64b5f6, stop:1 #2196f3);
                color: white; border: 2px solid #1976d2; border-radius: 8px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                           stop:0 #42a5f5, stop:1 #1976d2);
            }
        """)
        self.mic_button.setToolTip(self.tr("mic_tooltip_off"))
        if self.audio_stream:
            try:
                self.audio_stream.stop_stream()
                self.audio_stream.close()
            except:
                pass
            self.audio_stream = None

    def audio_listener_thread(self):
        try:
            p = pyaudio.PyAudio()
            input_device_index = None
            try:
                input_device_index = p.get_default_input_device_info()['index']
            except:
                for i in range(p.get_device_count()):
                    if p.get_device_info_by_index(i)['maxInputChannels'] > 0:
                        input_device_index = i
                        break
            if input_device_index is None:
                raise Exception("Aucun périphérique d'entrée audio trouvé")
            self.audio_stream = p.open(
                format=pyaudio.paInt16, channels=1, rate=16000,
                input=True, input_device_index=input_device_index,
                frames_per_buffer=8000)
            rec = vosk.KaldiRecognizer(self.vosk_model, 16000)
            rec.SetWords(True)
            while self.is_listening:
                try:
                    data = self.audio_stream.read(4000, exception_on_overflow=False)
                    if rec.AcceptWaveform(data):
                        result = json.loads(rec.Result())
                        text = result.get("text", "")
                        if text:
                            self.audio_queue.put(text)
                except Exception as read_error:
                    if self.is_listening:
                        continue
                    else:
                        break
        except Exception as e:
            self.audio_queue.put(("error", str(e)))
        finally:
            if self.audio_stream:
                try:
                    self.audio_stream.stop_stream()
                    self.audio_stream.close()
                except:
                    pass

    def process_audio_queue(self):
        try:
            while not self.audio_queue.empty():
                text = self.audio_queue.get_nowait()
                if isinstance(text, tuple) and text[0] == "error":
                    self.stop_listening()
                    return
                self.process_voice_input(text)
        except queue.Empty:
            pass

    def process_voice_input(self, text):
        number_words_fr = {
            "zéro": "0", "un": "1", "deux": "2", "trois": "3", "quatre": "4",
            "cinq": "5", "six": "6", "sept": "7", "huit": "8", "neuf": "9",
            "dix": "10", "onze": "11", "douze": "12", "treize": "13",
            "quatorze": "14", "quinze": "15", "seize": "16", "dix-sept": "17",
            "dix-huit": "18", "dix-neuf": "19", "vingt": "20", "trente": "30",
            "quarante": "40", "cinquante": "50", "soixante": "60",
            "soixante-dix": "70", "quatre-vingt": "80", "quatre-vingts": "80",
            "quatre-vingt-dix": "90"
        }
        number_words_ar = {
            "صفر": "0", "واحد": "1", "اثنان": "2", "ثلاثة": "3", "أربعة": "4",
            "خمسة": "5", "ستة": "6", "سبعة": "7", "ثمانية": "8", "تسعة": "9",
            "عشرة": "10", "عشرون": "20", "ثلاثون": "30", "أربعون": "40",
            "خمسون": "50", "ستون": "60", "سبعون": "70", "ثمانون": "80", "تسعون": "90"
        }
        number_words_en = {
            "zero": "0", "one": "1", "two": "2", "three": "3", "four": "4",
            "five": "5", "six": "6", "seven": "7", "eight": "8", "nine": "9",
            "ten": "10", "eleven": "11", "twelve": "12", "thirteen": "13",
            "fourteen": "14", "fifteen": "15", "sixteen": "16", "seventeen": "17",
            "eighteen": "18", "nineteen": "19", "twenty": "20", "thirty": "30",
            "forty": "40", "fifty": "50", "sixty": "60", "seventy": "70",
            "eighty": "80", "ninety": "90"
        }
        decimal_words_fr = {"vingt-cinq": "25", "vingt cinq": "25",
                             "cinquante": "5", "cinq": "5",
                             "soixante-quinze": "75", "soixante quinze": "75"}
        decimal_words_en = {"twenty-five": "25", "twenty five": "25",
                             "fifty": "5", "five": "5",
                             "seventy-five": "75", "seventy five": "75"}
        decimal_words_ar = {"ربع": "25", "نصف": "5", "ثلاثة أرباع": "75"}

        text = text.lower().strip()
        words = text.split()
        i = 0
        while i < len(words):
            word = words[i]
            if word.isdigit():
                for digit in word:
                    self.add_digit(digit)
                i += 1
            elif word in number_words_fr:
                for digit in number_words_fr[word]:
                    self.add_digit(digit)
                i += 1
            elif word in number_words_ar:
                for digit in number_words_ar[word]:
                    self.add_digit(digit)
                i += 1
            elif word in number_words_en:
                for digit in number_words_en[word]:
                    self.add_digit(digit)
                i += 1
            elif word in ["virgule", "point", "فاصل", "فاصلة"]:
                self.add_decimal_point()
                i += 1
                if i < len(words):
                    for n in [3, 2, 1]:
                        if i + n - 1 < len(words):
                            phrase = " ".join(words[i:i + n])
                            for dw in [decimal_words_fr, decimal_words_en, decimal_words_ar]:
                                if phrase in dw:
                                    for digit in dw[phrase]:
                                        self.add_digit(digit)
                                    i += n
                                    break
                            else:
                                continue
                            break
            elif word in ["valider", "entrer", "enter", "ok", "تأكيد"]:
                self.validate_number()
                i += 1
            elif word in ["effacer", "supprimer", "delete", "clear", "مسح"]:
                self.backspace()
                i += 1
            elif word in ["nouveau", "new", "جديد"]:
                self.clear_numbers()
                i += 1
            elif word in ["total", "المجموع", "مجموع"]:
                if self.current_number != "":
                    self.validate_number()
                self.history_of_totals.append(self.total)
                self.update_display()
                i += 1
            else:
                i += 1

    # ── Language change ──────────────────────────────────────────────────────

    def _refresh_all_ui_texts(self):
        """Update all UI text labels to the current language."""
        self.setWindowTitle(self.tr("title"))
        self.file_menu.setTitle(self.tr("file_menu"))
        self.language_menu.setTitle(self.tr("language_menu"))
        self.help_menu.setTitle(self.tr("help_menu"))
        self.shortcuts_action.setText(self.tr("shortcuts_action"))
        self.export_action.setText(self.tr("export"))
        self.stats_action.setText(self.tr("statistics"))
        self.about_action.setText(self.tr("about"))
        self.quit_action.setText(self.tr("quit"))
        self.undo_action.setText(self.tr("undo") + "    Ctrl+Z")
        self.redo_action.setText(self.tr("redo") + "    Ctrl+Y")
        self.arabic_tunisia_action.setText("    → " + self.tr("arabic_tunisia"))
        self.arabic_standard_action.setText("    → " + self.tr("arabic_standard"))
        self.current_label.setText(self.tr("current_input"))
        self.clear_numbers_button.setText(self.tr("clear_numbers"))
        self.clear_numbers_button.setToolTip(self.tr("tooltip_clear_numbers"))
        self.clear_totals_button.setText(self.tr("clear_totals"))
        self.clear_totals_button.setToolTip(self.tr("tooltip_clear_totals"))
        self.total_value_button.setToolTip(self.tr("tooltip_total"))
        self.enter_button.setToolTip(self.tr("tooltip_enter"))
        self.backspace_button.setToolTip(self.tr("tooltip_backspace"))
        if self.is_listening:
            self.mic_button.setText(self.tr("mic_on"))
            self.mic_button.setToolTip(self.tr("mic_tooltip_on"))
        else:
            self.mic_button.setText(self.tr("mic_off"))
            self.mic_button.setToolTip(self.tr("mic_tooltip_off"))
        self.historique_table.setHorizontalHeaderLabels(
            [self.tr("counter_header"), self.tr("value_header")])
        self.historique_table.setToolTip(self.tr("tooltip_history"))
        self.history_table.setHorizontalHeaderLabels(
            [self.tr("total_header"), self.tr("value_header")])
        self.history_table.setToolTip(self.tr("tooltip_totals"))

    def change_language(self, lang_code):
        self.current_language = lang_code
        self._refresh_all_ui_texts()
        if lang_code == "ar":
            self.setLayoutDirection(Qt.RightToLeft)
        else:
            self.setLayoutDirection(Qt.LeftToRight)

    # ── Core calculator logic ────────────────────────────────────────────────

    def add_digit(self, digit):
        self.current_number += str(digit)
        self.update_current_display()

    def add_and_validate_digit(self, digit):
        self.current_number += str(digit)
        self.validate_number()

    def add_decimal_point(self):
        if "." not in self.current_number:
            if self.current_number == "":
                self.current_number = "0."
            else:
                self.current_number += "."
            self.update_current_display()

    def validate_number(self):
        if self.current_number != "":
            self._push_undo()
            value = float(self.current_number)
            self.historique.append(value)
            self.total += value
            self.current_number = ""
            self.update_display()
            self.update_current_display()

    def add_fraction(self, fraction):
        self._push_undo()
        if self.current_number == "":
            value = fraction
        else:
            value = float(self.current_number) + fraction
        self.historique.append(value)
        self.total += value
        self.current_number = ""
        self.update_display()
        self.update_current_display()

    def backspace(self):
        """Delete: selected row (with confirmation) or last char/number."""
        # Check selection in number history
        selected_numbers = self.historique_table.selectedIndexes()
        if selected_numbers:
            row = selected_numbers[0].row()
            if 0 <= row < len(self.historique):
                reply = QMessageBox.question(
                    self,
                    self.tr("confirm_backspace_title"),
                    self.tr("confirm_backspace_row_msg"),
                    QMessageBox.Yes | QMessageBox.No,
                    QMessageBox.No
                )
                if reply == QMessageBox.Yes:
                    self._push_undo()
                    self.total -= self.historique[row]
                    del self.historique[row]
                    self.update_display()
                return

        # Check selection in totals history
        selected_totals = self.history_table.selectedIndexes()
        if selected_totals:
            row = selected_totals[0].row()
            if 0 <= row < len(self.history_of_totals):
                reply = QMessageBox.question(
                    self,
                    self.tr("confirm_backspace_title"),
                    self.tr("confirm_backspace_row_msg"),
                    QMessageBox.Yes | QMessageBox.No,
                    QMessageBox.No
                )
                if reply == QMessageBox.Yes:
                    self._push_undo()
                    del self.history_of_totals[row]
                    self.update_display()
                return

        # No selection: delete last char of current input
        if self.current_number != "":
            self.current_number = self.current_number[:-1]
            self.update_current_display()
        elif len(self.historique) > 0:
            self._push_undo()
            last_value = self.historique.pop()
            self.total -= last_value
            self.update_display()

    def update_current_display(self):
        self.current_input_label.setText(self.current_number)

    def update_display(self):
        self.historique_table.blockSignals(True)
        self.historique_table.setRowCount(len(self.historique))
        total_numbers = len(self.historique)
        for i, value in enumerate(self.historique):
            counter_item = QTableWidgetItem(f"{i + 1}/{total_numbers}")
            counter_item.setFlags(counter_item.flags() & ~Qt.ItemIsEditable)
            counter_item.setTextAlignment(Qt.AlignCenter)
            counter_item.setFont(QFont("Arial", 10, QFont.Bold))
            self.historique_table.setItem(i, 0, counter_item)
            value_item = QTableWidgetItem(f"{value:.2f}")
            value_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.historique_table.setItem(i, 1, value_item)
        self.historique_table.setColumnHidden(0, len(self.historique) == 0)
        self.historique_table.blockSignals(False)

        self.history_table.setRowCount(len(self.history_of_totals))
        total_totals = len(self.history_of_totals)
        for i, total in enumerate(self.history_of_totals):
            label_item = QTableWidgetItem(f"{self.tr('total_header')} {i + 1}/{total_totals}")
            label_item.setFlags(label_item.flags() & ~Qt.ItemIsEditable)
            label_item.setFont(QFont("Arial", 10, QFont.Bold))
            self.history_table.setItem(i, 0, label_item)
            value_item = QTableWidgetItem(f"{total:.2f}")
            value_item.setFlags(value_item.flags() & ~Qt.ItemIsEditable)
            value_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.history_table.setItem(i, 1, value_item)

        self.total_value_button.setText(f'{self.total:.2f}')

    def on_historique_changed(self, item):
        if item.column() == 1:
            row = item.row()
            try:
                old_value = self.historique[row]
                new_value = float(item.text().replace(',', '.'))
                self._push_undo()
                self.historique[row] = new_value
                self.total = self.total - old_value + new_value
                self.total_value_button.setText(f'{self.total:.2f}')
            except ValueError:
                self.update_display()

    def save_total_from_button(self):
        self._push_undo()
        if self.current_number != "":
            self.validate_number()
            # validate_number already pushed undo; pop the duplicate
            # Actually we want one unified undo for both actions.
            # Simplest: just let both pushes coexist (two undos).
        self.history_of_totals.append(self.total)
        self.update_display()

    def clear_numbers(self):
        """Clear all numbers — ask confirmation first."""
        reply = QMessageBox.question(
            self,
            self.tr("confirm_clear_numbers_title"),
            self.tr("confirm_clear_numbers_msg"),
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        if reply == QMessageBox.Yes:
            self._push_undo()
            self.historique = []
            self.current_number = ""
            self.total = 0
            self.update_display()
            self.update_current_display()

    def clear_totals(self):
        """Clear all totals — ask confirmation first."""
        reply = QMessageBox.question(
            self,
            self.tr("confirm_clear_totals_title"),
            self.tr("confirm_clear_totals_msg"),
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        if reply == QMessageBox.Yes:
            self._push_undo()
            self.history_of_totals = []
            self.update_display()

    def export_to_excel(self):
        filename, _ = QFileDialog.getSaveFileName(
            self, self.tr("save_dialog"), "", self.tr("excel_files"))
        if filename:
            try:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Historiques"
                ws['A1'] = self.tr("history_numbers_header")
                ws['A1'].font = ExcelFont(bold=True, size=12)
                ws['D1'] = self.tr("history_totals_header")
                ws['D1'].font = ExcelFont(bold=True, size=12)
                ws['A2'] = self.tr("counter_header")
                ws['B2'] = self.tr("value_header")
                ws['D2'] = self.tr("total_header")
                ws['E2'] = self.tr("value_header")
                for cell in ['A2', 'B2', 'D2', 'E2']:
                    ws[cell].font = ExcelFont(bold=True)
                    ws[cell].alignment = Alignment(horizontal='center')
                total_numbers = len(self.historique)
                for i, value in enumerate(self.historique, 1):
                    ws[f'A{i+2}'] = f"{i}/{total_numbers}"
                    ws[f'B{i+2}'] = value
                    ws[f'B{i+2}'].number_format = '0.00'
                total_totals = len(self.history_of_totals)
                for i, total in enumerate(self.history_of_totals, 1):
                    ws[f'D{i+2}'] = f"{self.tr('total_header')} {i}/{total_totals}"
                    ws[f'E{i+2}'] = total
                    ws[f'E{i+2}'].number_format = '0.00'
                ws.column_dimensions['A'].width = 12
                ws.column_dimensions['B'].width = 15
                ws.column_dimensions['D'].width = 15
                ws.column_dimensions['E'].width = 15
                wb.save(filename)
                QMessageBox.information(self, self.tr("export_success"),
                                        self.tr("export_success_msg").format(filename))
            except Exception as e:
                QMessageBox.critical(self, self.tr("export_error"),
                                     self.tr("export_error_msg").format(str(e)))

    def show_statistics(self):
        if len(self.history_of_totals) == 0:
            QMessageBox.information(self, self.tr("no_data"), self.tr("no_data_msg"))
            return
        dialog = QDialog(self)
        dialog.setWindowTitle(self.tr("stats_title"))
        dialog.setGeometry(100, 100, 900, 600)
        layout = QVBoxLayout()
        totals_array = np.array(self.history_of_totals)
        mean_value = np.mean(totals_array)
        std_value = np.std(totals_array)
        stats_label = QLabel(
            f"<b>{self.tr('mean')}:</b> {mean_value:.2f} | "
            f"<b>{self.tr('std_dev')}:</b> {std_value:.2f}", dialog)
        stats_label.setFont(QFont("Arial", 12))
        stats_label.setAlignment(Qt.AlignCenter)
        stats_label.setStyleSheet("padding: 10px; background-color: #f0f0f0;")
        layout.addWidget(stats_label)
        fig = Figure(figsize=(8, 5))
        canvas = FigureCanvas(fig)
        ax1 = fig.add_subplot(121)
        ax2 = fig.add_subplot(122)
        x_values = list(range(1, len(self.history_of_totals) + 1))
        ax1.plot(x_values, self.history_of_totals, 'bo-', linewidth=2, markersize=8,
                 label=self.tr("total_header"))
        ax1.axhline(y=mean_value, color='r', linestyle='--', linewidth=2,
                    label=f"{self.tr('mean')}: {mean_value:.2f}")
        ax1.fill_between(x_values, mean_value - std_value, mean_value + std_value,
                         alpha=0.2, color='red')
        ax1.set_xlabel(f"{self.tr('total_header')} #", fontsize=11)
        ax1.set_ylabel(self.tr("value_header"), fontsize=11)
        ax1.set_title(f"{self.tr('stats_title')}", fontsize=12, fontweight='bold')
        ax1.grid(True, alpha=0.3)
        ax1.legend()
        ax2.hist(self.history_of_totals, bins=min(10, len(self.history_of_totals)),
                 color='steelblue', edgecolor='black', alpha=0.7)
        ax2.axvline(x=mean_value, color='r', linestyle='--', linewidth=2,
                    label=f"{self.tr('mean')}: {mean_value:.2f}")
        ax2.set_xlabel(self.tr("value_header"), fontsize=11)
        ax2.set_ylabel("Fréquence", fontsize=11)
        ax2.set_title("Histogramme", fontsize=12, fontweight='bold')
        ax2.grid(True, alpha=0.3, axis='y')
        ax2.legend()
        fig.tight_layout()
        layout.addWidget(canvas)
        close_button = QPushButton("Fermer", dialog)
        close_button.clicked.connect(dialog.close)
        close_button.setFixedHeight(40)
        close_button.setFont(QFont("Arial", 11))
        layout.addWidget(close_button)
        dialog.setLayout(layout)
        dialog.exec_()

    def show_about(self):
        QMessageBox.about(
            self, self.tr("about_title"),
            f"<h3>{self.tr('title')}</h3>"
            f"<p>{self.tr('about_version')}</p>"
            f"<p>{self.tr('about_desc')}</p>"
            f"<p><b>{self.tr('about_shortcuts')}</b></p><ul>"
            f"<li>{self.tr('shortcut_digits')}</li>"
            f"<li>{self.tr('shortcut_decimal')}</li>"
            f"<li>{self.tr('shortcut_enter')}</li>"
            f"<li>{self.tr('shortcut_insert')}</li>"
            f"<li>{self.tr('shortcut_escape')}</li>"
            f"<li>{self.tr('shortcut_delete')}</li>"
            f"<li>{self.tr('shortcut_backspace')}</li>"
            f"<li>{self.tr('shortcut_undo')}</li>"
            f"<li>{self.tr('shortcut_redo')}</li>"
            f"</ul><p><b>{self.tr('about_voice')}</b></p><ul>"
            f"<li>{self.tr('voice_numbers')}</li>"
            f"<li>{self.tr('voice_decimal')}</li>"
            f"<li>{self.tr('voice_fractions')}</li>"
            f"<li>{self.tr('voice_validate')}</li>"
            f"<li>{self.tr('voice_delete')}</li>"
            f"<li>{self.tr('voice_new')}</li>"
            f"<li>{self.tr('voice_total')}</li>"
            f"</ul>"
        )

    # ── Keyboard handler ─────────────────────────────────────────────────────

    def keyPressEvent(self, event: QKeyEvent):
        key = event.key()
        modifiers = event.modifiers()

        # Ctrl+Z  → Undo
        if modifiers & Qt.ControlModifier and key == Qt.Key_Z:
            self.undo()
            return

        # Ctrl+Y  → Redo
        if modifiers & Qt.ControlModifier and key == Qt.Key_Y:
            self.redo()
            return

        if Qt.Key_0 <= key <= Qt.Key_9:
            if not (modifiers & Qt.ControlModifier):
                self.add_digit(key - Qt.Key_0)
        elif key in (Qt.Key_Period, Qt.Key_Comma):
            self.add_decimal_point()
        elif key in (Qt.Key_Return, Qt.Key_Enter):
            self.validate_number()
        elif key == Qt.Key_Insert or (modifiers & Qt.ControlModifier and key == Qt.Key_0):
            self._push_undo()
            if self.current_number != "":
                self.validate_number()
            self.history_of_totals.append(self.total)
            self.update_display()
        elif key == Qt.Key_Backspace:
            self.backspace()
        elif key == Qt.Key_Escape:
            self.clear_numbers()
        elif key == Qt.Key_Delete:
            self.clear_totals()


if __name__ == "__main__":
    app = QApplication(sys.argv)
    calc = Calculatrice()
    calc.show()
    sys.exit(app.exec_())